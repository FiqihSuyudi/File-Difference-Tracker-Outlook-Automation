# daily_email_gui.py
# Daily Comparison Email Builder (Classic Outlook Draft)
#
# Key updates:
# - Comparison + normalization matches the tracker wide export logic:
#     * normalize column names (strip)
#     * changed-cell detection compares ONLY columns present in BOTH files
#     * wide output: for changed rows, changed cells become "old → new"
#     * formatting: new-column headers blue, added rows green, deleted rows red,
#                   changed cells (only those containing "→") yellow
# - One consolidated "Excel (wide)" workbook with ONE sheet per selected job
# - No QA feature
#
# Requirements:
#   pip install pandas openpyxl matplotlib pywin32
#
# Run from:
#   C:\Users\muhammad.arafah\Documents\ID

import os
import re
import threading
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

try:
    import win32com.client  # type: ignore
except Exception:
    win32com = None


APP_TITLE = "Daily Comparison Email Builder"
DEFAULT_SOURCE_FOLDER = r"C:\Users\muhammad.arafah\Documents\ID"
OUT_DIRNAME = "_daily_compare_output"
IMG_DIRNAME = "images"
XLSX_DIRNAME = "reports"

# Excel color fills (match tracker)
FILL_YELLOW = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")
FILL_GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_BLUE   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

# Screenshot legend
LEGEND_ITEMS = [
    ("Changed cell", "#FFF3B0"),
    ("New row", "#C6EFCE"),
    ("Deleted row", "#FFC7CE"),
    ("New column (header)", "#BDD7EE"),
]


# -------------------- Utilities --------------------

def sizeof_fmt(num: int) -> str:
    for unit in ["B", "KB", "MB", "GB", "TB"]:
        if num < 1024 or unit == "TB":
            return f"{num:.1f} {unit}"
        num /= 1024


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    # tracker-style normalization: strip column names
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def read_any(path: str) -> pd.DataFrame:
    if path.lower().endswith(".csv"):
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path, engine="openpyxl")
    return normalize_columns(df)


def read_columns_only(path: str) -> List[str]:
    if not path or not os.path.exists(path):
        return []
    try:
        if path.lower().endswith(".csv"):
            df0 = pd.read_csv(path, nrows=0)
        else:
            df0 = pd.read_excel(path, engine="openpyxl", nrows=0)
        return [str(c).strip() for c in df0.columns]
    except Exception:
        return []


def safe_column_name(cols: List[str], key: str) -> Optional[str]:
    # allow case-insensitive matching
    if key in cols:
        return key
    kl = key.strip().lower()
    for c in cols:
        if str(c).strip().lower() == kl:
            return c
    return None


def file_info_line(path: str, df: pd.DataFrame) -> str:
    if not path:
        return "–"
    name = os.path.basename(path)
    rows = int(df.shape[0]) if df is not None else 0
    cols = int(df.shape[1]) if df is not None else 0
    try:
        size = sizeof_fmt(os.path.getsize(path))
    except Exception:
        size = "–"
    return f"{name} | {rows} rows, {cols} cols | {size}"


def find_latest_pair(folder: str, prefix: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Find latest 2 files matching:
      "<prefix>_YYYY-MM-DD.xlsx" or "<prefix> YYYY-MM-DD.xlsx" or csv
    Returns (old, new) where new is most recent.
    """
    p = Path(folder)
    if not p.exists():
        return None, None

    rx = re.compile(rf"^{re.escape(prefix)}[ _-](\d{{4}}-\d{{2}}-\d{{2}})\.(xlsx|xlsm|xls|csv)$", re.I)
    matches = []
    for f in p.iterdir():
        if not f.is_file():
            continue
        m = rx.match(f.name)
        if m:
            try:
                dt = datetime.strptime(m.group(1), "%Y-%m-%d")
                matches.append((dt, str(f)))
            except Exception:
                pass

    if len(matches) < 2:
        return None, None

    matches.sort(key=lambda x: x[0])
    return matches[-2][1], matches[-1][1]


def excel_safe_sheet_name(name: str, existing: set) -> str:
    # Excel: max 31 chars, no []:*?/\
    cleaned = re.sub(r"[\[\]\:\*\?\/\\]", " ", name).strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    base = cleaned[:31] if cleaned else "Sheet"
    out = base
    i = 2
    while out.lower() in existing:
        suffix = f" {i}"
        out = (base[: max(0, 31 - len(suffix))] + suffix).strip()
        i += 1
    existing.add(out.lower())
    return out


# -------------------- Tracker-matching Compare Logic (FULL + ADDED ONLY) --------------------

def compare_frames_tracker(
    df_old: pd.DataFrame,
    df_new: pd.DataFrame,
    key_cols: List[str],
    report_removed: bool = True,
    sparse_changed: bool = False,
) -> Tuple[pd.DataFrame, Dict, List[str]]:
    """
    Matches tracker v19 behavior (important bits):
      - compare changed cells ONLY on columns present in BOTH old and new
      - build WIDE changes-only dataset with __status__ and "old → new" strings for changed cells
      - union columns used for display/export (keys first)
    """
    df_old = normalize_columns(df_old)
    df_new = normalize_columns(df_new)

    # resolve key (case-insensitive) like your GUI dropdown might store
    resolved_keys = []
    for k in key_cols:
        ko = safe_column_name(list(df_old.columns), k)
        kn = safe_column_name(list(df_new.columns), k)
        if ko is None or kn is None:
            raise ValueError(f"Key '{k}' not present in both files.")
        # use NEW name (tracker assumes exact exists)
        resolved_keys.append(kn)
    key_cols = resolved_keys

    old_cols = set(df_old.columns)
    new_cols = set(df_new.columns)
    added_cols = sorted(list(new_cols - old_cols), key=lambda x: str(x).lower())
    removed_cols = sorted(list(old_cols - new_cols), key=lambda x: str(x).lower())

    # Union columns for export/display (keys first)
    all_cols_union = sorted(list(old_cols | new_cols), key=lambda x: (x not in key_cols, str(x).lower()))

    # Common columns for actual change detection
    common_cols = [c for c in all_cols_union if (c in old_cols and c in new_cols)]
    common_cols = key_cols + [c for c in common_cols if c not in key_cols]

    # Union indexed (for added/deleted rows)
    old_idxed_union = df_old.set_index(key_cols, drop=False).reindex(columns=all_cols_union)
    new_idxed_union = df_new.set_index(key_cols, drop=False).reindex(columns=all_cols_union)

    old_idx = old_idxed_union.index
    new_idx = new_idxed_union.index

    added_idx = new_idx.difference(old_idx)
    removed_idx = old_idx.difference(new_idx)
    common_idx = new_idx.intersection(old_idx)

    added_df = new_idxed_union.loc[added_idx] if len(added_idx) else pd.DataFrame(columns=all_cols_union)
    removed_df = old_idxed_union.loc[removed_idx] if len(removed_idx) else pd.DataFrame(columns=all_cols_union)

    # For change mask: compare only common columns
    old_common = old_idxed_union.loc[common_idx, common_cols]
    new_common = new_idxed_union.loc[common_idx, common_cols]

    old_common = old_common.sort_index()
    new_common = new_common.sort_index()
    old_common, new_common = old_common.align(new_common, join="inner", axis=None)

    if len(old_common) == 0:
        mask_changed = pd.DataFrame(False, index=old_common.index, columns=old_common.columns)
    else:
        mask_changed = (old_common != new_common) & ~(old_common.isna() & new_common.isna())

    changed_row_mask = mask_changed.any(axis=1)
    changed_pos = changed_row_mask.to_numpy().nonzero()[0]

    changed_cols = mask_changed.columns[mask_changed.any(axis=0)].tolist()
    changed_counts = mask_changed.sum(axis=0).astype(int).to_dict()
    changed_keys = list(mask_changed.index[mask_changed.any(axis=1)])

    records: List[Dict] = []

    # Added rows
    for _, row in added_df.iterrows():
        rec = row.to_dict()
        rec["__status__"] = "added"
        records.append(rec)

    # Removed rows
    if report_removed:
        for _, row in removed_df.iterrows():
            rec = row.to_dict()
            rec["__status__"] = "deleted"
            records.append(rec)

    # Changed rows
    for i in changed_pos:
        row_old_common = old_common.iloc[i]
        row_new_common = new_common.iloc[i]
        changed_cols_row = mask_changed.iloc[i]
        changed_cols_row = list(changed_cols_row[changed_cols_row].index)

        idx_key = old_common.index[i]
        row_new_union = new_idxed_union.loc[idx_key]

        if sparse_changed:
            rec = {k: row_new_union[k] for k in key_cols}
            for c in changed_cols_row:
                ov = row_old_common[c]
                nv = row_new_common[c]
                rec[c] = f"{ov} → {nv}"
        else:
            # tracker wide export uses sparse_changed=False (keep full row but replace changed cells with arrow)
            rec = row_new_union.to_dict()
            for c in changed_cols_row:
                ov = row_old_common[c]
                nv = row_new_common[c]
                rec[c] = f"{ov} → {nv}"

        rec["__status__"] = "changed"
        records.append(rec)

    # Ensure keys exist
    for rec in records:
        for k in key_cols:
            rec.setdefault(k, None)

    # stable column order (keys first, then encountered cols, then __status__)
    cols_union: List[str] = []
    for rec in records:
        for c in rec.keys():
            if c not in cols_union:
                cols_union.append(c)
    if "__status__" in cols_union:
        cols_union.remove("__status__")
    cols_union = key_cols + [c for c in cols_union if c not in key_cols] + ["__status__"]

    changes_only_df = pd.DataFrame(records, columns=cols_union)

    meta = {
        "key_cols": key_cols,
        "added_rows": int(len(added_idx)),
        "deleted_rows": int(len(removed_idx)),
        "changed_rows": int(changed_row_mask.sum()),
        "added_cols": added_cols,
        "removed_cols": removed_cols,
        "changed_cols": changed_cols,
        "changed_counts": changed_counts,
        "changed_keys": changed_keys,
        "cells_changed": int(mask_changed.sum().sum()) if not mask_changed.empty else 0,
        "identical": bool(
            len(added_idx) == 0 and len(removed_idx) == 0 and int(changed_row_mask.sum()) == 0
            and len(added_cols) == 0 and len(removed_cols) == 0
        ),
    }
    return changes_only_df, meta, all_cols_union


def compare_added_only_tracker(
    df_old: pd.DataFrame,
    df_new: pd.DataFrame,
    key_cols: List[str],
) -> Tuple[pd.DataFrame, Dict]:
    df_old = normalize_columns(df_old)
    df_new = normalize_columns(df_new)

    # resolve key case-insensitively
    resolved_keys = []
    for k in key_cols:
        ko = safe_column_name(list(df_old.columns), k)
        kn = safe_column_name(list(df_new.columns), k)
        if ko is None or kn is None:
            raise ValueError(f"Key '{k}' not present in both files.")
        resolved_keys.append(kn)
    key_cols = resolved_keys

    old_cols = set(df_old.columns)
    new_cols = set(df_new.columns)
    added_cols = sorted(list(new_cols - old_cols), key=lambda x: str(x).lower())

    old_idxed = df_old.set_index(key_cols, drop=False)
    new_idxed = df_new.set_index(key_cols, drop=False)

    added_idx = new_idxed.index.difference(old_idxed.index)
    added_df = new_idxed.loc[added_idx] if len(added_idx) else pd.DataFrame(columns=df_new.columns)

    records: List[Dict] = []
    for _, row in added_df.iterrows():
        rec = row.to_dict()
        rec["__status__"] = "added"
        records.append(rec)

    cols_union = list(df_new.columns) + ["__status__"]
    changes_only_df = pd.DataFrame(records, columns=cols_union) if records else pd.DataFrame(columns=cols_union)

    meta = {
        "key_cols": key_cols,
        "added_rows": int(len(added_idx)),
        "deleted_rows": 0,
        "changed_rows": 0,
        "added_cols": added_cols,
        "removed_cols": [],
        "changed_cols": [],
        "changed_counts": {},
        "changed_keys": [],
        "cells_changed": 0,
        "identical": bool(len(added_idx) == 0 and len(added_cols) == 0),
    }
    return changes_only_df, meta


# -------------------- Excel Export: Tracker "wide" formatting per sheet --------------------

def write_wide_sheet_like_tracker(
    wb: Workbook,
    sheet_name: str,
    changes_wide_df: pd.DataFrame,
    meta: Dict,
):
    ws = wb.create_sheet(sheet_name)

    headers = list(changes_wide_df.columns)
    ws.append(headers)

    # header format
    for j, c in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        if c in set(meta.get("added_cols", [])):
            cell.fill = FILL_BLUE

    # data rows
    for _, row in changes_wide_df.iterrows():
        ws.append([row.get(h) for h in headers])

    # styling by status (exact tracker rules)
    if "__status__" in headers:
        status_col = headers.index("__status__") + 1
        for i in range(2, ws.max_row + 1):
            st = ws.cell(row=i, column=status_col).value
            if st == "added":
                for j in range(1, status_col):
                    ws.cell(row=i, column=j).fill = FILL_GREEN
            elif st == "deleted":
                for j in range(1, status_col):
                    ws.cell(row=i, column=j).fill = FILL_RED
            elif st == "changed":
                for j in range(1, status_col):
                    val = ws.cell(row=i, column=j).value
                    if isinstance(val, str) and "→" in val:
                        ws.cell(row=i, column=j).fill = FILL_YELLOW

    # autosize columns (tracker-like, capped)
    for j in range(1, ws.max_column + 1):
        mx = 10
        for i in range(1, ws.max_row + 1):
            v = ws.cell(row=i, column=j).value
            if v is not None:
                mx = max(mx, len(str(v)))
        ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = min(mx + 2, 60)

    return ws


def export_consolidated_wide_workbook(
    out_path: str,
    per_job_payloads: List[Tuple[str, pd.DataFrame, Dict]],
):
    """
    per_job_payloads: list of (job_name, changes_wide_df, meta)
    Creates ONE workbook, one sheet per job, tracker-wide formatting.
    """
    wb = Workbook()
    # Remove default sheet safely (we'll add at least one sheet)
    default_ws = wb.active
    wb.remove(default_ws)

    existing = set()
    any_sheet = False

    for job_name, df_wide, meta in per_job_payloads:
        sheet_name = excel_safe_sheet_name(job_name, existing)

        # Even if empty, keep a minimal sheet so the workbook is valid + visible
        if df_wide is None or df_wide.empty:
            df_wide = pd.DataFrame(columns=(meta.get("key_cols", []) + ["__status__"]))
        write_wide_sheet_like_tracker(wb, sheet_name, df_wide, meta)
        any_sheet = True

    if not any_sheet:
        raise RuntimeError("No sheets to export.")

    # Ensure at least one visible sheet (openpyxl edge-cases)
    for ws in wb.worksheets:
        ws.sheet_state = "visible"
    wb.save(out_path)


# -------------------- Screenshot Rendering --------------------

def render_job_image(
    out_png: str,
    job_name: str,
    mode: str,
    key: str,
    old_path: str,
    new_path: str,
    df_old: pd.DataFrame,
    df_new: pd.DataFrame,
    preview_df: pd.DataFrame,
    meta: Dict,
    max_rows_preview: int = 12,
):
    # slightly wider for readability
    fig = plt.figure(figsize=(14.8, 5.2), dpi=120)
    fig.patch.set_facecolor("white")

    fig.text(0.02, 0.93, f"{job_name}:", fontsize=13, fontweight="bold")
    fig.text(0.02, 0.885, f"Key: {key}   |   Mode: {mode}", fontsize=10)

    old_info = file_info_line(old_path, df_old)
    new_info = file_info_line(new_path, df_new)
    fig.text(0.02, 0.845, f"OLD: {old_info}", fontsize=9)
    fig.text(0.02, 0.815, f"NEW: {new_info}", fontsize=9)

    if meta.get("identical"):
        fig.text(
            0.02, 0.775,
            "IDENTICAL — both files are the same (no added/deleted/changed rows or columns).",
            fontsize=10, fontweight="bold"
        )
    else:
        line = (
            f"Rows: +{meta.get('added_rows',0)}  "
            f"-{meta.get('deleted_rows',0)}  "
            f"changed_rows={meta.get('changed_rows',0)}   |   "
            f"Cells changed={meta.get('cells_changed',0)}   |   "
            f"Columns: +{len(meta.get('added_cols',[]))}  "
            f"-{len(meta.get('removed_cols',[]))}  "
            f"changed_cols={len(meta.get('changed_cols',[]))}"
        )
        fig.text(0.02, 0.775, line, fontsize=10)

    x = 0.02
    fig.text(x, 0.73, "Legend:", fontsize=9)
    x += 0.07
    for label, color in LEGEND_ITEMS:
        fig.text(
            x, 0.73, f"  {label}  ", fontsize=9,
            bbox=dict(facecolor=color, edgecolor="#999999", boxstyle="square,pad=0.2")
        )
        x += 0.16

    ax = fig.add_axes([0.02, 0.06, 0.96, 0.64])
    ax.axis("off")

    if preview_df is None or preview_df.empty:
        ax.text(0.0, 0.95, "No changes found.", fontsize=11)
        fig.savefig(out_png, bbox_inches="tight")
        plt.close(fig)
        return

    df_show = preview_df.head(max_rows_preview).copy()

    row_colors = []
    for _, r in df_show.iterrows():
        st = str(r.get("Status", "")).lower()
        if st == "added":
            row_colors.append("#C6EFCE")
        elif st == "deleted":
            row_colors.append("#FFC7CE")
        elif st == "changed":
            row_colors.append("#FFF9D7")
        else:
            row_colors.append("white")

    col_labels = list(df_show.columns)
    cell_text = df_show.values.tolist()

    table = ax.table(
        cellText=cell_text,
        colLabels=col_labels,
        loc="upper left",
        cellLoc="left"
    )
    table.auto_set_font_size(False)
    table.set_fontsize(8)
    table.scale(1.0, 1.25)

    # widen columns a bit for readability (best-effort)
    try:
        table.auto_set_column_width(col=list(range(len(col_labels))))
    except Exception:
        pass

    added_cols = set(meta.get("added_cols", []))
    for j, col in enumerate(col_labels):
        cell = table[(0, j)]
        cell.set_text_props(weight="bold")
        cell.set_facecolor("#BDD7EE" if col in added_cols else "#F2F2F2")

    for i in range(len(df_show)):
        for j in range(len(col_labels)):
            table[(i + 1, j)].set_facecolor(row_colors[i])

    fig.savefig(out_png, bbox_inches="tight")
    plt.close(fig)


# -------------------- Compact Preview (email table) from wide changes --------------------

def build_compact_preview_from_wide(
    changes_wide: pd.DataFrame,
    key_cols: List[str],
    max_summary_cols: int = 5,
) -> pd.DataFrame:
    """
    Compact preview similar to your email version:
    columns: key(s), Status, Column, Old, New
    """
    if changes_wide is None or changes_wide.empty:
        cols = key_cols + ["Status", "Column", "Old", "New"]
        return pd.DataFrame(columns=cols)

    status_col = "__status__"
    data_cols = [c for c in changes_wide.columns if c not in key_cols and c != status_col]

    records = []
    for _, row in changes_wide.iterrows():
        status = row.get(status_col, "")
        key_vals = {k: row.get(k) for k in key_cols}

        if status == "changed":
            for c in data_cols:
                val = row.get(c)
                if pd.isna(val):
                    continue
                s = str(val)
                if "→" not in s:
                    continue
                old, new = s.split("→", 1)
                records.append({
                    **key_vals,
                    "Status": "changed",
                    "Column": c,
                    "Old": old.strip(),
                    "New": new.strip(),
                })

        elif status in ("added", "deleted"):
            parts = []
            for c in data_cols[:max_summary_cols]:
                v = row.get(c)
                if pd.isna(v) or v is None or str(v).strip() == "":
                    continue
                parts.append(f"{c}={v}")
            summary = "; ".join(parts) if parts else "(no values)"
            if status == "added":
                records.append({**key_vals, "Status": "added", "Column": "(row)", "Old": "", "New": summary})
            else:
                records.append({**key_vals, "Status": "deleted", "Column": "(row)", "Old": summary, "New": ""})

    preview = pd.DataFrame(records) if records else pd.DataFrame(columns=key_cols + ["Status", "Column", "Old", "New"])
    cols_order = key_cols + ["Status", "Column", "Old", "New"]
    for c in cols_order:
        if c not in preview.columns:
            preview[c] = ""
    return preview[cols_order]


# -------------------- Outlook Draft --------------------

def create_outlook_draft(
    to: str,
    cc: str,
    subject: str,
    html_body: str,
    inline_images: List[str],
    attachments: List[str],
):
    if win32com is None:
        raise RuntimeError("pywin32 (win32com.client) not available.")

    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)  # olMailItem

    if to.strip():
        mail.To = to.strip()
    if cc.strip():
        mail.CC = cc.strip()

    mail.Subject = subject

    # Attach images inline using CID
    cids = []
    for idx, img_path in enumerate(inline_images, start=1):
        cid = f"img{idx}"
        attach = mail.Attachments.Add(img_path)
        attach.PropertyAccessor.SetProperty(
            "http://schemas.microsoft.com/mapi/proptag/0x3712001F",
            cid
        )
        cids.append(cid)

    body = html_body
    for idx, cid in enumerate(cids, start=1):
        body = body.replace(f"{{{{CID:img{idx}}}}}", cid)

    mail.HTMLBody = body

    # Additional attachments (excel consolidated)
    for p in attachments:
        mail.Attachments.Add(p)

    mail.Save()
    mail.Display()
    return True


# -------------------- Jobs Model --------------------

@dataclass
class JobConfig:
    name: str
    prefix: str
    default_key: str
    default_mode: str  # "full" or "added_only"


DEFAULT_JOBS = [
    JobConfig("Urlaubsplanung", "Urlaubsplanung ID", "ID", "full"),
    JobConfig("Bestellungen Detail", "Bestellungen Detail ID", "UNIQUE", "full"),
    JobConfig("Bestellungen", "Bestellungen ID", "ID", "full"),
    JobConfig("Mitarbeiterübersicht", "Mitarbeiterübersicht ID", "ID", "full"),
    JobConfig("Projektübersicht", "Projektübersicht ID", "ID", "full"),
    JobConfig("Lokationen", "Lokationen ID", "ID", "full"),
    JobConfig("Projektplanung", "Projektplanung ID", "ID", "added_only"),
]


# -------------------- GUI --------------------

class ScrollableFrame(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        canvas = tk.Canvas(self, highlightthickness=0)
        vsb = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        self.inner = ttk.Frame(canvas)

        self.inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.inner, anchor="nw")
        canvas.configure(yscrollcommand=vsb.set)

        canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        def _on_enter(_):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
        def _on_leave(_):
            canvas.unbind_all("<MouseWheel>")
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")

        canvas.bind("<Enter>", _on_enter)
        canvas.bind("<Leave>", _on_leave)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1200x740")
        self.minsize(1050, 660)

        self.source_folder = tk.StringVar(value=DEFAULT_SOURCE_FOLDER)
        self.to_var = tk.StringVar(value="")
        self.cc_var = tk.StringVar(value="")

        self.progress = tk.DoubleVar(value=0.0)
        self.progress_text = tk.StringVar(value="Ready.")

        self.jobs_ui: List[Dict] = []

        self._build_ui()
        self._build_jobs(DEFAULT_JOBS)

    # ----- Thread-safe UI helpers -----

    def ui_log(self, msg: str):
        def _do():
            ts = datetime.now().strftime("%H:%M:%S")
            self.log.insert("end", f"[{ts}] {msg}\n")
            self.log.see("end")
            self.update_idletasks()
        self.after(0, _do)

    def ui_progress(self, pct: float, msg: str):
        def _do():
            self.progress.set(max(0, min(100, pct)))
            self.progress_text.set(msg)
            self.update_idletasks()
        self.after(0, _do)

    def ui_warning(self, title: str, msg: str):
        self.after(0, lambda: messagebox.showwarning(title, msg))

    def ui_error(self, title: str, msg: str):
        self.after(0, lambda: messagebox.showerror(title, msg))

    # ----- UI -----

    def _build_ui(self):
        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=10)

        ttk.Label(top, text="Source folder:").pack(side="left")
        ttk.Entry(top, textvariable=self.source_folder, width=70).pack(side="left", padx=6)
        ttk.Button(top, text="Change…", command=self.pick_folder).pack(side="left", padx=6)

        row2 = ttk.Frame(self)
        row2.pack(fill="x", padx=10, pady=(0, 8))
        ttk.Label(row2, text="To:").pack(side="left")
        ttk.Entry(row2, textvariable=self.to_var, width=45).pack(side="left", padx=6)
        ttk.Label(row2, text="CC:").pack(side="left", padx=(20, 0))
        ttk.Entry(row2, textvariable=self.cc_var, width=45).pack(side="left", padx=6)

        self.scroll = ScrollableFrame(self)
        self.scroll.pack(fill="both", expand=True, padx=10, pady=5)

        bottom = ttk.Frame(self)
        bottom.pack(fill="x", padx=10, pady=8)

        ttk.Progressbar(bottom, variable=self.progress, maximum=100).pack(
            side="left", fill="x", expand=True, padx=(0, 10)
        )
        ttk.Label(bottom, textvariable=self.progress_text, width=55, anchor="w").pack(side="left")

        ttk.Button(bottom, text="Auto-detect all latest", command=self.auto_detect_all_clicked).pack(side="left", padx=8)
        ttk.Button(bottom, text="Create Outlook draft", command=self.create_draft_clicked).pack(side="right")

        log_frame = ttk.LabelFrame(self, text="Steps / Logs")
        log_frame.pack(fill="both", expand=False, padx=10, pady=(0, 10))
        self.log = tk.Text(log_frame, height=9, wrap="word")
        self.log.pack(fill="both", expand=True, padx=6, pady=6)

    def pick_folder(self):
        p = filedialog.askdirectory(title="Select source folder", initialdir=self.source_folder.get())
        if p:
            self.source_folder.set(p)

    def _build_jobs(self, jobs: List[JobConfig]):
        for child in self.scroll.inner.winfo_children():
            child.destroy()
        self.jobs_ui.clear()

        for job in jobs:
            self._add_job_row(job)

    def _add_job_row(self, job: JobConfig):
        row = ttk.Frame(self.scroll.inner)
        row.pack(fill="x", pady=6)

        enabled_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(row, variable=enabled_var).pack(side="left", padx=(0, 8))

        ttk.Label(row, text=job.name, width=18).pack(side="left")

        ttk.Label(row, text="Key:").pack(side="left", padx=(8, 2))
        key_var = tk.StringVar(value=job.default_key)
        key_combo = ttk.Combobox(row, textvariable=key_var, width=18, state="readonly")
        key_combo["values"] = [job.default_key]
        key_combo.pack(side="left", padx=(0, 10))

        ttk.Label(row, text="Mode:").pack(side="left", padx=(8, 2))
        mode_var = tk.StringVar(value=job.default_mode)
        mode_combo = ttk.Combobox(row, textvariable=mode_var, width=12, state="readonly", values=["full", "added_only"])
        mode_combo.pack(side="left", padx=(0, 10))

        # consolidated wide export include checkbox (per job)
        include_wide_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(row, text="Include in 1 Excel (wide)", variable=include_wide_var).pack(side="left", padx=(8, 10))

        old_path = tk.StringVar(value="")
        new_path = tk.StringVar(value="")

        ttk.Button(row, text="Pick OLD", command=lambda: self.pick_file(old_path, key_combo)).pack(side="left", padx=4)
        ttk.Button(row, text="Pick NEW", command=lambda: self.pick_file(new_path, key_combo)).pack(side="left", padx=4)
        ttk.Button(
            row,
            text="Auto-detect latest",
            command=lambda: self.auto_detect_one(job.prefix, old_path, new_path, key_combo)
        ).pack(side="left", padx=6)

        line = ttk.Frame(self.scroll.inner)
        line.pack(fill="x")
        lbl = ttk.Label(line, text="", anchor="w")
        lbl.pack(side="left", fill="x", expand=True)

        def refresh_paths_label(*_):
            lbl.config(text=f"OLD: {old_path.get()}    NEW: {new_path.get()}")

        old_path.trace_add("write", refresh_paths_label)
        new_path.trace_add("write", refresh_paths_label)
        refresh_paths_label()

        self.jobs_ui.append({
            "job": job,
            "enabled": enabled_var,
            "key_var": key_var,
            "key_combo": key_combo,
            "mode_var": mode_var,
            "include_wide_var": include_wide_var,
            "old_path": old_path,
            "new_path": new_path,
        })

    def pick_file(self, var: tk.StringVar, key_combo: ttk.Combobox):
        p = filedialog.askopenfilename(
            title="Select file",
            filetypes=[("Excel/CSV", "*.xlsx;*.xls;*.xlsm;*.csv"), ("All files", "*.*")]
        )
        if p:
            var.set(p)
            self.populate_key_dropdown_fast(key_combo)

    def populate_key_dropdown_fast(self, key_combo: ttk.Combobox):
        owner = None
        for ui in self.jobs_ui:
            if ui["key_combo"] == key_combo:
                owner = ui
                break
        if owner is None:
            return

        old_path = owner["old_path"].get()
        new_path = owner["new_path"].get()
        if not old_path or not new_path:
            return

        cols_old = read_columns_only(old_path)
        cols_new = read_columns_only(new_path)
        if not cols_new:
            return

        map_new = {c.lower(): c for c in cols_new}
        inter = []
        for c in cols_old:
            cl = c.lower()
            if cl in map_new:
                inter.append(map_new[cl])

        values = inter if inter else cols_new
        current = owner["key_var"].get().strip()

        best = None
        if current:
            for v in values:
                if v.lower() == current.lower():
                    best = v
                    break
        if best:
            owner["key_var"].set(best)

        key_combo["values"] = values

    def auto_detect_one(self, prefix: str, old_var: tk.StringVar, new_var: tk.StringVar, key_combo: ttk.Combobox):
        folder = self.source_folder.get().strip()
        old_p, new_p = find_latest_pair(folder, prefix)
        if old_p and new_p:
            old_var.set(old_p)
            new_var.set(new_p)
            self.populate_key_dropdown_fast(key_combo)
        else:
            messagebox.showwarning("Not found", f"Could not find at least 2 files for prefix:\n{prefix}")

    def auto_detect_all_clicked(self):
        t = threading.Thread(target=self._auto_detect_all_worker, daemon=True)
        t.start()

    def _auto_detect_all_worker(self):
        try:
            self.ui_progress(0, "Auto-detecting latest files…")
            self.ui_log("Auto-detect all latest started…")

            folder = self.source_folder.get().strip()
            if not folder:
                self.ui_warning("Folder missing", "Please set the source folder.")
                return

            total = len(self.jobs_ui)
            for i, ui in enumerate(self.jobs_ui, start=1):
                job: JobConfig = ui["job"]
                prefix = job.prefix

                self.ui_progress((i-1)/total*100, f"[{i}/{total}] Finding {job.name}…")
                self.ui_log(f"[{i}/{total}] Finding latest pair: {job.name} ({prefix})")

                old_p, new_p = find_latest_pair(folder, prefix)
                if not old_p or not new_p:
                    self.ui_log(f"WARNING: Not enough files found for {job.name}")
                    continue

                def apply_paths(u=ui, o=old_p, n=new_p):
                    u["old_path"].set(o)
                    u["new_path"].set(n)
                    self.populate_key_dropdown_fast(u["key_combo"])

                self.after(0, apply_paths)
                self.ui_log(f"Selected OLD={os.path.basename(old_p)} | NEW={os.path.basename(new_p)}")
                self.ui_progress(i/total*100, f"[{i}/{total}] Done {job.name}")

            self.ui_progress(100, "Auto-detect finished.")
            self.ui_log("Auto-detect all latest finished.")
        except Exception as e:
            self.ui_log(f"FATAL in auto-detect: {e}")
            self.ui_error("Auto-detect error", str(e))

    def create_draft_clicked(self):
        if win32com is None:
            messagebox.showerror(
                "Missing dependency",
                "win32com.client not available.\n\nInstall:\n  pip install pywin32\n\nRestart VS Code after install."
            )
            return
        t = threading.Thread(target=self._create_draft_worker, daemon=True)
        t.start()

    def _create_draft_worker(self):
        try:
            self.ui_progress(0, "Starting…")
            self.ui_log("Starting draft creation…")

            folder = self.source_folder.get().strip()
            if not folder:
                self.ui_warning("Folder missing", "Please set the source folder.")
                return

            selected = [ui for ui in self.jobs_ui if ui["enabled"].get()]
            if not selected:
                self.ui_warning("No jobs", "Please enable at least one job.")
                return

            out_root = Path(folder) / OUT_DIRNAME
            img_dir = out_root / IMG_DIRNAME
            xlsx_dir = out_root / XLSX_DIRNAME
            img_dir.mkdir(parents=True, exist_ok=True)
            xlsx_dir.mkdir(parents=True, exist_ok=True)

            all_images: List[str] = []
            body_parts: List[str] = []
            wide_payloads: List[Tuple[str, pd.DataFrame, Dict]] = []

            total = len(selected)
            today = datetime.now().strftime("%Y-%m-%d")
            ts_run = datetime.now().strftime("%Y%m%d_%H%M%S")

            for idx, ui in enumerate(selected, start=1):
                job: JobConfig = ui["job"]
                key = ui["key_var"].get().strip()
                mode = ui["mode_var"].get().strip()
                old_path = ui["old_path"].get().strip()
                new_path = ui["new_path"].get().strip()
                include_wide = ui["include_wide_var"].get()

                self.ui_progress((idx - 1) / total * 100, f"[{idx}/{total}] Validating {job.name}…")
                self.ui_log(f"[{idx}/{total}] Job: {job.name} | mode={mode} | key={key}")

                if not old_path or not new_path:
                    self.ui_log(f"ERROR: Missing OLD/NEW for {job.name}")
                    continue
                if not os.path.exists(old_path) or not os.path.exists(new_path):
                    self.ui_log(f"ERROR: File not found for {job.name}")
                    continue

                self.ui_progress((idx - 1) / total * 100 + 4, f"[{idx}/{total}] Reading files…")
                self.ui_log(f"Reading OLD: {os.path.basename(old_path)}")
                df_old = read_any(old_path)
                self.ui_log(f"Reading NEW: {os.path.basename(new_path)}")
                df_new = read_any(new_path)

                self.ui_progress((idx - 1) / total * 100 + 10, f"[{idx}/{total}] Comparing…")
                try:
                    if mode == "added_only":
                        changes_wide, meta = compare_added_only_tracker(df_old, df_new, [key])
                    else:
                        changes_wide, meta, _ = compare_frames_tracker(
                            df_old, df_new, [key], report_removed=True, sparse_changed=False
                        )
                except Exception as e:
                    self.ui_log(f"ERROR comparing {job.name}: {e}")
                    continue

                # Compact preview for screenshot/email
                preview = build_compact_preview_from_wide(changes_wide, meta.get("key_cols", [key]))

                self.ui_progress((idx - 1) / total * 100 + 20, f"[{idx}/{total}] Rendering screenshot…")
                png_path = str(img_dir / f"{job.name}_{today}_{ts_run}_{idx}.png")

                render_job_image(
                    out_png=png_path,
                    job_name=job.name,
                    mode=mode,
                    key=key,
                    old_path=old_path,
                    new_path=new_path,
                    df_old=df_old,
                    df_new=df_new,
                    preview_df=preview,
                    meta=meta,
                    max_rows_preview=12
                )

                all_images.append(png_path)
                cid_placeholder = f"{{{{CID:img{len(all_images)}}}}}"
                body_parts.append(f"""
                    <div style="margin-bottom:18px;">
                      <img src="cid:{cid_placeholder}" style="max-width:100%; border:1px solid #ddd;">
                    </div>
                """)
                self.ui_log(f"Rendered image: {png_path}")

                if include_wide:
                    self.ui_log(f"Queue wide sheet for consolidated workbook: {job.name}")
                    wide_payloads.append((job.name, changes_wide, meta))

                self.ui_progress(idx / total * 100, f"[{idx}/{total}] Done {job.name}")

            if not all_images:
                self.ui_warning("No output", "No images were created. Check logs for errors.")
                return

            attachments: List[str] = []

            # Build ONE consolidated wide workbook if any job checked
            if wide_payloads:
                self.ui_progress(92, "Building consolidated wide workbook…")
                out_xlsx = str(xlsx_dir / f"DailyCompare_wide_{today}_{ts_run}.xlsx")
                try:
                    export_consolidated_wide_workbook(out_xlsx, wide_payloads)
                    attachments.append(out_xlsx)
                    try:
                        size_mb = os.path.getsize(out_xlsx) / (1024 * 1024)
                        self.ui_log(f"Saved consolidated workbook: {out_xlsx} ({size_mb:.1f} MB)")
                    except Exception:
                        self.ui_log(f"Saved consolidated workbook: {out_xlsx}")
                except Exception as e:
                    self.ui_log(f"WARNING: Failed to build consolidated workbook: {e}")

            subject = f"Daily Comparison Result — {today}"
            html = f"""
            <html>
            <body style="font-family:Calibri,Arial,sans-serif;">
              <h2 style="margin:0 0 10px 0;">Daily Comparison Result ({today})</h2>
              <p style="margin:0 0 12px 0;">Auto-generated draft. Please review before sending.</p>
              {''.join(body_parts)}
              <hr>
              <p style="font-size:11px; color:#666;">
                Generated by Daily Comparison Email Builder.
              </p>
            </body>
            </html>
            """

            self.ui_progress(98, "Creating Outlook draft…")
            self.ui_log("Creating Outlook draft in Classic Outlook…")

            create_outlook_draft(
                to=self.to_var.get(),
                cc=self.cc_var.get(),
                subject=subject,
                html_body=html,
                inline_images=all_images,
                attachments=attachments
            )

            self.ui_progress(100, "Done — Draft created.")
            self.ui_log("DONE: Outlook Draft created and opened.")
        except Exception as e:
            self.ui_log(f"FATAL: {e}")
            self.ui_error("Error", str(e))


if __name__ == "__main__":
    App().mainloop()
